#!/usr/bin/env python3
"""
Build a SYSTEMATIC, EXHAUSTIVE investable-asset taxonomy covering (as completely
as is practical) every asset class and sub-class in the world, and write it to
an Excel workbook (multi-sheet) plus a flat master CSV.

Hierarchy columns:
  Tier            -> broad grouping bucket
  Asset Class     -> top-level class
  Sub-Class       -> next level down
  Sector/Strategy -> sector, strategy, structural slice, industry, or region
  Sub-Sector      -> finest grain (industry, sub-industry, tranche, contract, pair)
  Geography       -> Domestic (US) / Developed ex-US / EM / Frontier / Global
  Examples        -> representative tickers / instruments (illustrative, not advice)
  Typ. Yield      -> rough income range where meaningful
  Risk Tier       -> 1 (capital-preservation) .. 6 (max beta / first loss)
  Liquidity       -> Intraday / Daily / Periodic / Illiquid
  Notes           -> what you're really betting on / key risk

Equity sectors use the full GICS (2023) classification, which is GLOBAL.
Yields/risk tiers are rough illustrative ranges, NOT advice. Tickers anchor a
category; they are not recommendations.
"""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

COLUMNS = [
    "Tier", "Asset Class", "Sub-Class", "Sector/Strategy", "Sub-Sector",
    "Geography", "Examples", "Typ. Yield", "Risk Tier", "Liquidity", "Notes",
]

ROWS = []
def R(*args):
    # pad to 11 columns
    args = list(args) + [""] * (11 - len(args))
    ROWS.append(tuple(args[:11]))

# =====================================================================
# TIER 0 — CASH & CASH EQUIVALENTS
# =====================================================================
T = "Cash & Equivalents"
R(T,"Cash","Bank deposits","Checking/Savings","FDIC savings","Domestic","HYSA","4-5%",1,"Daily","Insured to limits; reinvestment risk")
R(T,"Cash","Bank deposits","Certificates of deposit","Brokered CDs","Domestic","-","4-5%",1,"Periodic","Term lockup; FDIC insured")
R(T,"Cash","Bank deposits","Sweep/cash management","Bank sweep","Domestic","-","3-5%",1,"Daily","Brokerage sweep accounts")
R(T,"Cash","Money market","Govt money market funds","T-bill/repo MMF","Domestic","SGOV,BIL,SPAXX","4-5%",1,"Daily","Near-zero credit risk; rate-sensitive")
R(T,"Cash","Money market","Prime money market funds","CP/CD MMF","Domestic","-","4-5%",1,"Daily","Slight credit spread over govt")
R(T,"Cash","Money market","Tax-exempt MMF","Muni MMF","Domestic","-","3%TE",1,"Daily","Tax-exempt cash")
R(T,"Cash","T-bills","Ultra-short Treasuries","0-3m bills","Domestic","BIL,SGOV","4-5%",1,"Intraday","Risk-free rate proxy")
R(T,"Cash","Commercial paper","Short corp paper","CP","Domestic","-","4-5%",1,"Periodic","Short-term corporate funding")
R(T,"Cash","Repo","Repurchase agreements","Tri-party repo","Domestic","-","4-5%",1,"Daily","Collateralized overnight lending")
R(T,"Cash","Stablecoins","Fiat-backed","USD stablecoins","Global","USDC,USDT","0-5%*",2,"Intraday","Peg/issuer risk; *yield via lending")
R(T,"Cash","Foreign cash","FX deposits","Non-USD deposits","Global","-","varies",2,"Daily","FX exposure")

# =====================================================================
# TIER 1 — GOVERNMENT FIXED INCOME
# =====================================================================
T = "Govt Fixed Income"
R(T,"US Treasuries","Nominal","Short duration","1-3y","Domestic","SHY,VGSH","4-5%",1,"Intraday","Duration/rate risk only")
R(T,"US Treasuries","Nominal","Intermediate","3-10y","Domestic","IEF,VGIT","4%",2,"Intraday","Belly of the curve")
R(T,"US Treasuries","Nominal","Long duration","10-30y","Domestic","TLT,VGLT,EDV","4-5%",3,"Intraday","High rate convexity; big drawdowns")
R(T,"US Treasuries","Inflation-linked","TIPS","Short/Long TIPS","Domestic","TIP,VTIP,SCHP,LTPZ","2%+CPI",2,"Intraday","Bets on realized inflation")
R(T,"US Treasuries","Floating rate","FRN","2y FRN","Domestic","USFR,TFLO","4-5%",1,"Intraday","Minimal duration; tracks bills")
R(T,"US Treasuries","STRIPS","Zero coupon","Long zeros","Domestic","GOVZ,EDV","4-5%",4,"Intraday","Max duration convexity")
R(T,"US Treasuries","Savings bonds","Series I/EE","I-bonds","Domestic","-","CPI",1,"Periodic","Direct-from-Treasury; lockups")
R(T,"Agency","GSE debt","Agency debentures","FNMA/FHLB/FHLMC","Domestic","-","4-5%",1,"Daily","Implicit govt backing")
R(T,"Agency","Supranational","Multilateral dev banks","World Bank/IBRD/ADB","Global","-","4-5%",1,"Daily","AAA supranational issuers")
R(T,"Sovereign ex-US","Developed","DM govt bonds","Bunds/JGBs/Gilts","Developed ex-US","BWX,IGOV","1-4%",2,"Daily","FX + foreign rate risk")
R(T,"Sovereign ex-US","Developed FX-hedged","Hedged DM govt","Hedged global agg","Developed ex-US","BNDX","2-3%",2,"Daily","Strips FX, keeps rate diversification")
R(T,"Sovereign ex-US","Global linkers","Foreign inflation-linked","Intl TIPS","Global","WIP,GTIP","CPI+",2,"Daily","Global inflation protection")
R(T,"Sovereign ex-US","EM hard ccy","USD sovereign","EM USD bonds","EM","EMB,PCY","6-8%",4,"Daily","Spread + default risk, no FX")
R(T,"Sovereign ex-US","EM local ccy","Local sovereign","EM local bonds","EM","EMLC,LEMB","6-8%",4,"Daily","FX + local rate + default")
R(T,"Sovereign ex-US","Frontier","Frontier sovereign","Frontier USD bonds","Frontier","-","8-12%",5,"Periodic","Illiquid, high default/political risk")

# =====================================================================
# TIER 2 — MUNICIPAL
# =====================================================================
T = "Municipal"
R(T,"Munis","General obligation","Investment-grade GO","National GO","Domestic","MUB,VTEB","3-4%TE",2,"Daily","Tax-exempt; state credit risk")
R(T,"Munis","Revenue","IG revenue","Toll/utility/hospital","Domestic","-","3-4%TE",2,"Daily","Project cash-flow backed")
R(T,"Munis","High yield","HY munis","Non-rated/spec muni","Domestic","HYD,HYMB","4-5%TE",3,"Daily","Lower-rated issuers; default risk")
R(T,"Munis","State-specific","Single-state","CA/NY munis","Domestic","-","3-4%TE",2,"Daily","Double tax-exempt for residents")
R(T,"Munis","Taxable munis","BABs","Build America Bonds","Domestic","BAB","4-5%",2,"Daily","Taxable; for non-taxable accounts")
R(T,"Munis","Pre-refunded","Escrowed","Pre-re munis","Domestic","-","3%TE",1,"Daily","Treasury-collateralized; very safe")
R(T,"Munis","Short munis","0-5y muni","Short tax-exempt","Domestic","SUB,SHM","2-3%TE",1,"Daily","Low duration tax-exempt")
R(T,"Munis","Tobacco/special tax","Settlement-backed","Tobacco bonds","Domestic","-","4-5%TE",4,"Daily","Settlement revenue risk")

# =====================================================================
# TIER 3 — CORPORATE CREDIT (incl. rating buckets)
# =====================================================================
T = "Corporate Credit"
R(T,"Investment grade","Broad IG","Aggregate corp","US IG corp","Domestic","LQD,VCIT","4-5%",2,"Intraday","Spread + rate risk")
R(T,"Investment grade","Short IG","1-5y IG","Short corp","Domestic","IGSB,VCSH","4-5%",2,"Intraday","Low duration credit")
R(T,"Investment grade","Long IG","10y+ IG","Long corp","Domestic","IGLB,VCLT","5%",3,"Intraday","Duration + spread")
R(T,"Investment grade","AAA/AA","Highest grade","AAA-AA corp","Domestic","QLTA","4-5%",2,"Daily","Top-quality corporates")
R(T,"Investment grade","A-rated","Single-A","A corp","Domestic","-","4-5%",2,"Daily","Upper-medium grade")
R(T,"Investment grade","Crossover BBB","BBB","BBB-tilt","Domestic","-","5%",3,"Intraday","Fallen-angel risk")
R(T,"High yield","Broad HY","BB/B/CCC","US HY","Domestic","HYG,JNK,USHY","6-8%",4,"Intraday","Default + spread risk")
R(T,"High yield","BB-rated","Upper HY","BB corp","Domestic","-","6-7%",4,"Daily","Higher-quality junk")
R(T,"High yield","B/CCC-rated","Lower HY","B/CCC corp","Domestic","-","8-10%",5,"Daily","Elevated default risk")
R(T,"High yield","Fallen angels","Ex-IG downgrades","Fallen angel HY","Domestic","FALN","6-7%",4,"Daily","Higher quality HY tilt")
R(T,"High yield","Short HY","0-5y HY","Short HY","Domestic","SHYG,SJNK","6-7%",4,"Intraday","Lower duration HY")
R(T,"High yield","Distressed","CCC/defaulted","Distressed debt","Domestic","-","10%+",6,"Illiquid","Workout/recovery bet")
R(T,"Convertibles","Convertible bonds","Balanced converts","US converts","Domestic","CWB,ICVT","2-4%",4,"Daily","Equity upside + bond floor")
R(T,"Green/ESG bonds","Labeled bonds","Green/social/sustainability","Use-of-proceeds bonds","Global","-","4-5%",2,"Daily","ESG-labeled credit")
R(T,"International corp","DM corporate","IG/HY foreign","Euro/UK corp","Developed ex-US","IBND,HYXU","3-6%",3,"Daily","FX + foreign credit")
R(T,"International corp","EM corporate","EM USD corp","EM corp bonds","EM","CEMB,EMCB","6-8%",4,"Daily","EM corporate credit")

# =====================================================================
# TIER 4 — SECURITIZED / STRUCTURED CREDIT  (CLO ladder lives here)
# =====================================================================
T = "Securitized Credit"
R(T,"Agency MBS","Pass-throughs","30y/15y MBS","Agency RMBS","Domestic","MBB,VMBS","4-5%",2,"Intraday","Prepayment/convexity risk")
R(T,"Agency MBS","Agency CMBS","Multifamily","GNMA/FNMA CMBS","Domestic","-","4-5%",2,"Daily","Govt-backed commercial")
R(T,"Non-agency MBS","RMBS","Prime/Alt-A/legacy","Non-agency RMBS","Domestic","-","5-7%",4,"Periodic","Housing credit risk")
R(T,"Non-agency MBS","Whole loans","Residential whole loans","Whole-loan pools","Domestic","-","6-8%",4,"Illiquid","Direct mortgage credit")
R(T,"CMBS","Commercial MBS","Conduit/SASB","CMBS senior","Domestic","CMBS*","5-7%",4,"Daily","CRE credit risk")
R(T,"ABS","Consumer ABS","Auto/card/student","Consumer ABS","Domestic","-","5-6%",3,"Daily","Consumer credit cycle")
R(T,"ABS","Esoteric ABS","Aircraft/royalty/data","Esoteric ABS","Domestic","-","6-8%",4,"Periodic","Niche collateral risk")
R(T,"ABS","SBA loans","Govt-guaranteed","SBA pools","Domestic","-","5-6%",2,"Daily","Small-business, govt-backed")
R(T,"CLO","CLO debt","AAA tranche","Top tranche","Domestic","JAAA,AAA,CLOA","5-6%",1,"Daily","Spread normalize; near-zero default")
R(T,"CLO","CLO debt","AA tranche","Senior mezz","Domestic","-","6%",2,"Daily","Slightly subordinated, floating")
R(T,"CLO","CLO debt","A tranche","Mezz","Domestic","-","6-7%",2,"Daily","Floating-rate mezzanine")
R(T,"CLO","CLO debt","BBB tranche","Mezz","Domestic","JBBB,CLOZ","6-7%",2,"Daily","Spread tightening + floating rate")
R(T,"CLO","CLO debt","BB tranche","Lower mezz","Domestic","CLOZ","8-9%",4,"Daily","First-loss-adjacent; higher beta")
R(T,"CLO","CLO equity","Residual/equity","First-loss","Domestic","ECC,OXLC","15-20%+",6,"Daily","Everything-goes-right residual")
R(T,"CDO/other","Synthetic/other structured","CDO/CRT/CFO","Structured tranches","Domestic","-","varies",5,"Periodic","Tranche-specific risk")
R(T,"Multi-sector","Active securitized","Go-anywhere structured","Multi-sector credit","Domestic","-","6-8%",4,"Daily","Manager rotates structured sleeves")

# =====================================================================
# TIER 5 — BANK / SENIOR LOANS
# =====================================================================
T = "Loans"
R(T,"Senior loans","Leveraged loans","Broadly syndicated","Senior secured","Domestic","BKLN,SRLN","8%",4,"Daily","Floating; single-name default risk")
R(T,"Senior loans","Active loans","CLO-style mgmt","Bank loan funds","Domestic","FFRHX,SRLN","8%",4,"Daily","No structural leverage")
R(T,"Senior loans","Middle-market loans","Direct loans","MM senior","Domestic","-","9-11%",5,"Periodic","Smaller borrowers, less liquid")
R(T,"Consumer loans","Marketplace lending","P2P/fintech loans","Consumer credit","Domestic","-","6-9%",5,"Periodic","Unsecured consumer default risk")
R(T,"Trade finance","Receivables/factoring","Short-term trade","Trade finance","Global","-","6-9%",4,"Periodic","Counterparty/settlement risk")

# =====================================================================
# TIER 6 — PREFERRED & HYBRID CAPITAL
# =====================================================================
T = "Preferred & Hybrid"
R(T,"Preferreds","Traditional preferreds","$25 par perpetual","Bank/utility pfd","Domestic","PFF,PGX","6-7%",3,"Daily","Rate + credit; below senior debt")
R(T,"Preferreds","Institutional preferreds","$1000 par","Inst'l pfd","Domestic","PFFA,PFFD","6-7%",3,"Daily","Higher quality issuers")
R(T,"Preferreds","CLO-CEF preferreds","Term preferred","Senior to CEF common","Domestic","OXLCN,OXLCI,OXLCZ,ECC-D","7-8.5%",3,"Daily","Redeemed at $25 ahead of equity")
R(T,"Baby bonds","Exchange-traded debt","$25 baby bonds","Senior to pfd","Domestic","-","7-8%",3,"Daily","Bond claim, exchange-traded")
R(T,"Trust preferreds","TruPS","Bank trust pfd","Hybrid capital","Domestic","-","6-8%",3,"Daily","Deferrable hybrid")
R(T,"CoCos","Contingent convertibles","AT1 bank capital","CoCo bonds","Developed ex-US","-","7-9%",5,"Daily","Write-down/conversion trigger risk")
R(T,"Convertible pfd","Convertible preferreds","Equity-linked pfd","Convert pfd","Domestic","-","5-7%",4,"Daily","Hybrid equity upside")

# =====================================================================
# TIER 7 — PRIVATE CREDIT / DIRECT LENDING
# =====================================================================
T = "Private Credit"
R(T,"BDC","Listed BDC","Top-tier direct lender","Senior direct lending","Domestic","ARCC,MAIN,OBDC","9-10%",5,"Daily","Direct-lending equity; soft marks")
R(T,"BDC","BDC basket","Diversified BDC","BDC index","Domestic","BIZD,PBDC","9-11%",5,"Daily","Private-credit beta")
R(T,"BDC","Venture/specialty BDC","Venture debt","Tech lending BDC","Domestic","HTGC,TRIN","10-12%",5,"Daily","Higher-risk borrowers")
R(T,"Interval funds","Private credit interval","Non-traded","Direct lending interval","Domestic","-","8-10%",5,"Periodic","Quarterly liquidity gates")
R(T,"Private debt funds","LP direct lending","Drawdown funds","Senior/unitranche","Domestic","-","9-12%",5,"Illiquid","Lockup; J-curve")
R(T,"Private debt funds","Distressed/special sits","Opportunistic","Distressed credit","Global","-","12-15%",6,"Illiquid","Restructuring upside")
R(T,"Mezzanine","Mezz/PIK debt","Subordinated","Mezzanine","Domestic","-","11-14%",6,"Illiquid","Below senior; equity kickers")
R(T,"Real estate debt","Bridge/CRE loans","Private RE lending","CRE debt","Domestic","-","8-11%",5,"Illiquid","Property-secured private loans")
R(T,"Asset-based lending","ABL","Receivables/inventory","ABL","Domestic","-","9-12%",5,"Illiquid","Collateralized specialty finance")

# =====================================================================
# TIER 8 — US/SIZE/STYLE/FACTOR EQUITY EXPOSURES
# =====================================================================
T = "Equity Size/Style/Factor"
R(T,"Broad market","Total market","Cap-weighted","US total","Domestic","VTI,ITOT","1-2%",4,"Intraday","US equity beta")
R(T,"Broad market","Large cap","S&P 500","Large blend","Domestic","SPY,VOO,IVV","1-2%",4,"Intraday","Mega/large cap beta")
R(T,"Broad market","Mega cap","Top 50/100","Mega blend","Domestic","MGC,OEF","1%",4,"Intraday","Largest companies")
R(T,"Broad market","Mid cap","S&P 400","Mid blend","Domestic","IJH,VO","1-2%",4,"Intraday","Mid-cap premium")
R(T,"Broad market","Small cap","Russell 2000","Small blend","Domestic","IWM,VB,IJR","1%",5,"Intraday","Small-cap/economic beta")
R(T,"Broad market","Micro cap","Micro-cap index","Micro","Domestic","IWC","1%",5,"Daily","Illiquidity premium")
R(T,"Broad market","Equal weight","S&P 500 EW","Equal-weight","Domestic","RSP","1-2%",4,"Intraday","De-concentrated beta")
R(T,"Style","Growth","Large growth","Growth tilt","Domestic","VUG,IWF,QQQ","0-1%",4,"Intraday","Duration/long-growth bet")
R(T,"Style","Value","Large value","Value tilt","Domestic","VTV,IWD","2-3%",4,"Intraday","Cheapness factor")
R(T,"Style","Small value","Small-cap value","SCV","Domestic","VBR,AVUV","1-2%",5,"Intraday","Classic value+size premium")
R(T,"Style","Dividend growth","Quality dividends","Div growers","Domestic","SCHD,VIG,DGRO","2-3%",3,"Intraday","Quality income compounding")
R(T,"Style","High dividend","Yield tilt","High div","Domestic","HDV,SPYD,VYM","3-4%",3,"Intraday","Income-tilted equity")
R(T,"Style","Buyback","Shareholder yield","Buyback tilt","Domestic","PKW,SYLD","1-2%",4,"Intraday","Capital-return focus")
R(T,"Factor","Quality","High ROE/low debt","Quality factor","Domestic","QUAL","1-2%",4,"Intraday","Profitability factor")
R(T,"Factor","Momentum","Trend","Momentum factor","Domestic","MTUM","1%",4,"Intraday","Price persistence")
R(T,"Factor","Low volatility","Min vol","Low-vol factor","Domestic","USMV,SPLV","2%",3,"Intraday","Defensive factor")
R(T,"Factor","Multi-factor","Combined","Multi-factor","Domestic","LRGF,GSLC","1-2%",4,"Intraday","Diversified factor exposure")
R(T,"Factor","Profitability/yield","Shareholder yield","Multi-metric","Domestic","-","2-3%",4,"Intraday","Combined fundamentals")

# =====================================================================
# TIER 9 — GLOBAL EQUITY SECTORS (FULL GICS 2023 — applies globally)
# =====================================================================
T = "Global Equity Sectors (GICS)"
GICS = {
 "Energy": {
   "Energy Equipment & Services": {"Energy Equipment & Services": ["Oil & Gas Drilling","Oil & Gas Equipment & Services"]},
   "Oil, Gas & Consumable Fuels": {"Oil, Gas & Consumable Fuels": ["Integrated Oil & Gas","Oil & Gas Exploration & Production","Oil & Gas Refining & Marketing","Oil & Gas Storage & Transportation","Coal & Consumable Fuels"]},
 },
 "Materials": {
   "Materials": {
     "Chemicals": ["Commodity Chemicals","Diversified Chemicals","Fertilizers & Agricultural Chemicals","Industrial Gases","Specialty Chemicals"],
     "Construction Materials": ["Construction Materials"],
     "Containers & Packaging": ["Metal, Glass & Plastic Containers","Paper & Plastic Packaging Products & Materials"],
     "Metals & Mining": ["Aluminum","Diversified Metals & Mining","Copper","Gold","Precious Metals & Minerals","Silver","Steel"],
     "Paper & Forest Products": ["Forest Products","Paper Products"],
   },
 },
 "Industrials": {
   "Capital Goods": {
     "Aerospace & Defense": ["Aerospace & Defense"],
     "Building Products": ["Building Products"],
     "Construction & Engineering": ["Construction & Engineering"],
     "Electrical Equipment": ["Electrical Components & Equipment","Heavy Electrical Equipment"],
     "Industrial Conglomerates": ["Industrial Conglomerates"],
     "Machinery": ["Construction Machinery & Heavy Transportation Equipment","Agricultural & Farm Machinery","Industrial Machinery & Supplies & Components"],
     "Trading Companies & Distributors": ["Trading Companies & Distributors"],
   },
   "Commercial & Professional Services": {
     "Commercial Services & Supplies": ["Commercial Printing","Environmental & Facilities Services","Office Services & Supplies","Diversified Support Services","Security & Alarm Services"],
     "Professional Services": ["Human Resource & Employment Services","Research & Consulting Services","Data Processing & Outsourced Services"],
   },
   "Transportation": {
     "Air Freight & Logistics": ["Air Freight & Logistics"],
     "Passenger Airlines": ["Passenger Airlines"],
     "Marine Transportation": ["Marine Transportation"],
     "Ground Transportation": ["Rail Transportation","Cargo Ground Transportation","Passenger Ground Transportation"],
     "Transportation Infrastructure": ["Airport Services","Highways & Railtracks","Marine Ports & Services"],
   },
 },
 "Consumer Discretionary": {
   "Automobiles & Components": {
     "Automobile Components": ["Automotive Parts & Equipment","Tires & Rubber"],
     "Automobiles": ["Automobile Manufacturers","Motorcycle Manufacturers"],
   },
   "Consumer Durables & Apparel": {
     "Household Durables": ["Consumer Electronics","Home Furnishings","Homebuilding","Household Appliances","Housewares & Specialties"],
     "Leisure Products": ["Leisure Products"],
     "Textiles, Apparel & Luxury Goods": ["Apparel, Accessories & Luxury Goods","Footwear","Textiles"],
   },
   "Consumer Services": {
     "Hotels, Restaurants & Leisure": ["Casinos & Gaming","Hotels, Resorts & Cruise Lines","Leisure Facilities","Restaurants"],
     "Diversified Consumer Services": ["Education Services","Specialized Consumer Services"],
   },
   "Consumer Discretionary Distribution & Retail": {
     "Distributors": ["Distributors"],
     "Broadline Retail": ["Broadline Retail"],
     "Specialty Retail": ["Apparel Retail","Computer & Electronics Retail","Home Improvement Retail","Specialty Stores","Automotive Retail","Homefurnishing Retail"],
   },
 },
 "Consumer Staples": {
   "Consumer Staples Distribution & Retail": {
     "Consumer Staples Distribution & Retail": ["Drug Retail","Food Distributors","Food Retail","Consumer Staples Merchandise Retail"],
   },
   "Food, Beverage & Tobacco": {
     "Beverages": ["Brewers","Distillers & Vintners","Soft Drinks & Non-alcoholic Beverages"],
     "Food Products": ["Agricultural Products & Services","Packaged Foods & Meats"],
     "Tobacco": ["Tobacco"],
   },
   "Household & Personal Products": {
     "Household Products": ["Household Products"],
     "Personal Care Products": ["Personal Care Products"],
   },
 },
 "Health Care": {
   "Health Care Equipment & Services": {
     "Health Care Equipment & Supplies": ["Health Care Equipment","Health Care Supplies"],
     "Health Care Providers & Services": ["Health Care Distributors","Health Care Services","Health Care Facilities","Managed Health Care"],
     "Health Care Technology": ["Health Care Technology"],
   },
   "Pharmaceuticals, Biotechnology & Life Sciences": {
     "Biotechnology": ["Biotechnology"],
     "Pharmaceuticals": ["Pharmaceuticals"],
     "Life Sciences Tools & Services": ["Life Sciences Tools & Services"],
   },
 },
 "Financials": {
   "Banks": {"Banks": ["Diversified Banks","Regional Banks"]},
   "Financial Services": {
     "Financial Services": ["Diversified Financial Services","Multi-Sector Holdings","Specialized Finance","Commercial & Residential Mortgage Finance","Transaction & Payment Processing Services"],
     "Consumer Finance": ["Consumer Finance"],
   },
   "Insurance": {
     "Insurance": ["Insurance Brokers","Life & Health Insurance","Multi-line Insurance","Property & Casualty Insurance","Reinsurance"],
   },
 },
 "Information Technology": {
   "Software & Services": {
     "IT Services": ["IT Consulting & Other Services","Internet Services & Infrastructure"],
     "Software": ["Application Software","Systems Software"],
   },
   "Technology Hardware & Equipment": {
     "Communications Equipment": ["Communications Equipment"],
     "Technology Hardware, Storage & Peripherals": ["Technology Hardware, Storage & Peripherals"],
     "Electronic Equipment, Instruments & Components": ["Electronic Equipment & Instruments","Electronic Components","Electronic Manufacturing Services","Technology Distributors"],
   },
   "Semiconductors & Semiconductor Equipment": {
     "Semiconductors & Semiconductor Equipment": ["Semiconductor Materials & Equipment","Semiconductors"],
   },
 },
 "Communication Services": {
   "Telecommunication Services": {
     "Diversified Telecommunication Services": ["Alternative Carriers","Integrated Telecommunication Services"],
     "Wireless Telecommunication Services": ["Wireless Telecommunication Services"],
   },
   "Media & Entertainment": {
     "Media": ["Advertising","Broadcasting","Cable & Satellite","Publishing"],
     "Entertainment": ["Movies & Entertainment","Interactive Home Entertainment"],
     "Interactive Media & Services": ["Interactive Media & Services"],
   },
 },
 "Utilities": {
   "Utilities": {
     "Electric Utilities": ["Electric Utilities"],
     "Gas Utilities": ["Gas Utilities"],
     "Multi-Utilities": ["Multi-Utilities"],
     "Water Utilities": ["Water Utilities"],
     "Independent Power & Renewable Electricity Producers": ["Independent Power Producers & Energy Traders","Renewable Electricity"],
   },
 },
 "Real Estate": {
   "Equity Real Estate Investment Trusts (REITs)": {
     "Equity REITs": ["Diversified REITs","Industrial REITs","Hotel & Resort REITs","Office REITs","Health Care REITs","Multi-Family Residential REITs","Single-Family Residential REITs","Retail REITs","Other Specialized REITs","Self-Storage REITs","Telecom Tower REITs","Timber REITs","Data Center REITs"],
   },
   "Real Estate Management & Development": {
     "Real Estate Mgmt & Development": ["Diversified Real Estate Activities","Real Estate Operating Companies","Real Estate Development","Real Estate Services"],
   },
 },
}
# Sector-level SPDR fallbacks
SECTOR_ETF = {
 "Energy":"XLE","Materials":"XLB","Industrials":"XLI","Consumer Discretionary":"XLY",
 "Consumer Staples":"XLP","Health Care":"XLV","Financials":"XLF","Information Technology":"XLK",
 "Communication Services":"XLC","Utilities":"XLU","Real Estate":"XLRE",
}
# Best pure-play / closest-fit liquid ETF per GICS sub-industry (else sector SPDR).
# "~" prefix = sector/thematic proxy, not a pure play.
GICS_TICKER = {
 "Oil & Gas Drilling":"OIH","Oil & Gas Equipment & Services":"OIH,XES",
 "Integrated Oil & Gas":"IXC","Oil & Gas Exploration & Production":"XOP",
 "Oil & Gas Refining & Marketing":"CRAK","Oil & Gas Storage & Transportation":"AMLP,MLPX",
 "Coal & Consumable Fuels":"~XLE",
 "Commodity Chemicals":"~XLB","Diversified Chemicals":"~XLB","Fertilizers & Agricultural Chemicals":"MOO,VEGI",
 "Industrial Gases":"~XLB","Specialty Chemicals":"~XLB","Construction Materials":"PKB",
 "Metal, Glass & Plastic Containers":"~XLB","Paper & Plastic Packaging Products & Materials":"~XLB",
 "Aluminum":"~PICK","Diversified Metals & Mining":"XME,PICK","Copper":"COPX","Gold":"GDX,GDXJ",
 "Precious Metals & Minerals":"GDX","Silver":"SIL","Steel":"SLX",
 "Forest Products":"WOOD,CUT","Paper Products":"WOOD",
 "Aerospace & Defense":"ITA,PPA,XAR","Building Products":"PKB","Construction & Engineering":"PAVE",
 "Electrical Components & Equipment":"~XLI","Heavy Electrical Equipment":"~XLI","Industrial Conglomerates":"~XLI",
 "Construction Machinery & Heavy Transportation Equipment":"~XLI","Agricultural & Farm Machinery":"MOO",
 "Industrial Machinery & Supplies & Components":"~XLI","Trading Companies & Distributors":"~XLI",
 "Commercial Printing":"~XLI","Environmental & Facilities Services":"EVX","Office Services & Supplies":"~XLI",
 "Diversified Support Services":"~XLI","Security & Alarm Services":"~XLI",
 "Human Resource & Employment Services":"~XLI","Research & Consulting Services":"~XLI",
 "Data Processing & Outsourced Services":"IPAY,FINX",
 "Air Freight & Logistics":"IYT","Passenger Airlines":"JETS","Marine Transportation":"BOAT,SEA",
 "Rail Transportation":"IYT","Cargo Ground Transportation":"IYT","Passenger Ground Transportation":"~IYT",
 "Airport Services":"~IGF","Highways & Railtracks":"~IGF","Marine Ports & Services":"~IGF",
 "Automotive Parts & Equipment":"CARZ","Tires & Rubber":"~CARZ","Automobile Manufacturers":"CARZ,DRIV",
 "Motorcycle Manufacturers":"~CARZ","Consumer Electronics":"~XLY","Home Furnishings":"XHB",
 "Homebuilding":"ITB,XHB","Household Appliances":"~XHB","Housewares & Specialties":"~XLY",
 "Leisure Products":"PEJ","Apparel, Accessories & Luxury Goods":"~XLY","Footwear":"~XLY","Textiles":"~XLY",
 "Casinos & Gaming":"BJK,BETZ","Hotels, Resorts & Cruise Lines":"AWAY,PEJ","Leisure Facilities":"PEJ",
 "Restaurants":"EATZ","Education Services":"~XLY","Specialized Consumer Services":"~XLY",
 "Distributors":"~XLY","Broadline Retail":"XRT,RTH","Apparel Retail":"XRT",
 "Computer & Electronics Retail":"XRT","Home Improvement Retail":"XHB","Specialty Stores":"XRT",
 "Automotive Retail":"~CARZ","Homefurnishing Retail":"XHB",
 "Drug Retail":"~XLP","Food Distributors":"PBJ","Food Retail":"PBJ","Consumer Staples Merchandise Retail":"~XLP",
 "Brewers":"PBJ","Distillers & Vintners":"PBJ","Soft Drinks & Non-alcoholic Beverages":"PBJ",
 "Agricultural Products & Services":"MOO,PBJ","Packaged Foods & Meats":"PBJ","Tobacco":"~XLP",
 "Household Products":"~XLP","Personal Care Products":"~XLP",
 "Health Care Equipment":"IHI,XHE","Health Care Supplies":"IHI","Health Care Distributors":"IHF",
 "Health Care Services":"IHF","Health Care Facilities":"IHF","Managed Health Care":"IHF",
 "Health Care Technology":"EDOC,IHF","Biotechnology":"XBI,IBB","Pharmaceuticals":"PJP,IHE",
 "Life Sciences Tools & Services":"~XLV",
 "Diversified Banks":"KBE","Regional Banks":"KRE","Diversified Financial Services":"~XLF",
 "Multi-Sector Holdings":"~XLF","Specialized Finance":"~XLF","Commercial & Residential Mortgage Finance":"REM",
 "Transaction & Payment Processing Services":"IPAY,FINX","Consumer Finance":"~XLF",
 "Asset Management & Custody Banks":"IAI","Investment Banking & Brokerage":"IAI",
 "Diversified Capital Markets":"IAI","Financial Exchanges & Data":"IAI",
 "Insurance Brokers":"KIE,IAK","Life & Health Insurance":"KIE","Multi-line Insurance":"KIE",
 "Property & Casualty Insurance":"KBWP,KIE","Reinsurance":"~KIE",
 "IT Consulting & Other Services":"~IGV","Internet Services & Infrastructure":"FDN",
 "Application Software":"IGV","Systems Software":"IGV","Communications Equipment":"IYZ",
 "Technology Hardware, Storage & Peripherals":"~XLK","Electronic Equipment & Instruments":"~XLK",
 "Electronic Components":"~XLK","Electronic Manufacturing Services":"~XLK","Technology Distributors":"~XLK",
 "Semiconductor Materials & Equipment":"SOXX,SMH","Semiconductors":"SOXX,SMH",
 "Alternative Carriers":"IYZ","Integrated Telecommunication Services":"IYZ,VOX",
 "Wireless Telecommunication Services":"IYZ,VOX","Advertising":"~XLC","Broadcasting":"~XLC",
 "Cable & Satellite":"~XLC","Publishing":"~XLC","Movies & Entertainment":"PEJ",
 "Interactive Home Entertainment":"ESPO,HERO","Interactive Media & Services":"FCOM,XLC",
 "Electric Utilities":"XLU,IDU","Gas Utilities":"~XLU","Multi-Utilities":"~XLU",
 "Water Utilities":"PHO,FIW","Independent Power Producers & Energy Traders":"~XLU",
 "Renewable Electricity":"ICLN,TAN",
 "Diversified REITs":"VNQ","Industrial REITs":"INDS","Hotel & Resort REITs":"~VNQ","Office REITs":"~VNQ",
 "Health Care REITs":"~VNQ","Multi-Family Residential REITs":"REZ","Single-Family Residential REITs":"REZ",
 "Retail REITs":"~VNQ","Other Specialized REITs":"~VNQ","Self-Storage REITs":"~VNQ",
 "Telecom Tower REITs":"~VNQ","Timber REITs":"WOOD,CUT","Data Center REITs":"DTCR,SRVR",
 "Diversified Real Estate Activities":"~VNQ","Real Estate Operating Companies":"~VNQ",
 "Real Estate Development":"~VNQ","Real Estate Services":"~VNQ",
}
for sector, groups in GICS.items():
    for group, industries in groups.items():
        for industry, subs in industries.items():
            for sub in subs:
                ex = GICS_TICKER.get(sub) or ("~" + SECTOR_ETF.get(sector, "-"))
                note = "GICS sub-industry; pure-play ETF" if not ex.startswith("~") else "GICS sub-industry; sector/thematic proxy (no pure-play ETF)"
                R(T, sector, group, industry, sub, "Global", ex, "varies", 4, "Intraday", note)

# =====================================================================
# TIER 10 — SINGLE-COUNTRY & REGIONAL EQUITY
# =====================================================================
T = "Country/Regional Equity"
# Broad/regional
for nm, sub, geo, ex, risk in [
  ("Developed ex-US","EAFE","Developed ex-US","VEA,IEFA,EFA",4),
  ("Developed ex-US","DM small cap","Developed ex-US","SCZ,GWX",5),
  ("Developed ex-US","DM FX-hedged","Developed ex-US","HEFA,HEDJ",4),
  ("Europe","Europe broad","Developed ex-US","VGK,FEZ",4),
  ("Asia-Pacific","Asia ex-Japan","Global","AAXJ,VPL",5),
  ("Emerging markets","EM broad","EM","VWO,IEMG,EEM",5),
  ("Emerging markets","EM ex-China","EM","EMXC",5),
  ("Emerging markets","EM small cap","EM","EWX,DGS",5),
  ("Frontier markets","Frontier broad","Frontier","FM,FRN",6),
  ("Global","All-world","Global","ACWI,VT",4),
  ("Global","Global ex-US","Global","VXUS,CWI",4),
  ("Global","Global dividend","Global","IDV,VIGI,VYMI",3),
]:
    R(T,"Region/aggregate",nm,sub,"-",geo,ex,"2-4%",risk,"Intraday","Regional equity beta")
# Single countries: (Country, Geo, Example, Risk)
COUNTRIES = [
  ("United States","Domestic","SPY,VTI",4),
  ("Canada","Developed ex-US","EWC",4),
  ("United Kingdom","Developed ex-US","EWU,FLGB",4),
  ("Germany","Developed ex-US","EWG",4),
  ("France","Developed ex-US","EWQ",4),
  ("Italy","Developed ex-US","EWI",4),
  ("Spain","Developed ex-US","EWP",4),
  ("Netherlands","Developed ex-US","EWN",4),
  ("Switzerland","Developed ex-US","EWL",4),
  ("Sweden","Developed ex-US","EWD",4),
  ("Norway","Developed ex-US","ENOR",4),
  ("Denmark","Developed ex-US","EDEN",4),
  ("Finland","Developed ex-US","EFNL",4),
  ("Belgium","Developed ex-US","EWK",4),
  ("Austria","Developed ex-US","EWO",4),
  ("Ireland","Developed ex-US","EIRL",4),
  ("Portugal","Developed ex-US","PGAL",4),
  ("Japan","Developed ex-US","EWJ,DXJ,BBJP",4),
  ("Australia","Developed ex-US","EWA",4),
  ("New Zealand","Developed ex-US","ENZL",4),
  ("Hong Kong","Developed ex-US","EWH",4),
  ("Singapore","Developed ex-US","EWS",4),
  ("Israel","Developed ex-US","EIS",4),
  ("China","EM","MCHI,FXI,KWEB,ASHR",5),
  ("India","EM","INDA,EPI,SMIN",5),
  ("Taiwan","EM","EWT",5),
  ("South Korea","EM","EWY",5),
  ("Brazil","EM","EWZ,EWZS",5),
  ("Mexico","EM","EWW",5),
  ("Chile","EM","ECH",5),
  ("Colombia","EM","GXG",5),
  ("Peru","EM","EPU",5),
  ("Argentina","EM","ARGT",6),
  ("South Africa","EM","EZA",5),
  ("Saudi Arabia","EM","KSA",5),
  ("United Arab Emirates","EM","UAE",5),
  ("Qatar","EM","QAT",5),
  ("Kuwait","EM","KWT*",5),
  ("Turkey","EM","TUR",6),
  ("Poland","EM","EPOL",5),
  ("Greece","EM","GREK",5),
  ("Malaysia","EM","EWM",5),
  ("Thailand","EM","THD",5),
  ("Indonesia","EM","EIDO",5),
  ("Philippines","EM","EPHE",5),
  ("Vietnam","Frontier","VNM",6),
  ("Nigeria","Frontier","NGE",6),
  ("Egypt","Frontier","EGPT",6),
  ("Pakistan","Frontier","PAK",6),
  ("Bangladesh","Frontier","-",6),
]
for country, geo, ex, risk in COUNTRIES:
    R(T,"Single country",country,"Country equity","-",geo,ex,"1-5%",risk,"Intraday","Country-specific macro/FX/political risk")

# =====================================================================
# TIER 11 — THEMATIC EQUITIES
# =====================================================================
T = "Thematic Equities"
for sub, sec, ex, risk in [
  ("Technology themes","Artificial intelligence","BOTZ,ROBO,IRBO,AIQ",6),
  ("Technology themes","Cybersecurity","CIBR,HACK,BUG",6),
  ("Technology themes","Cloud/SaaS","SKYY,WCLD",6),
  ("Technology themes","Semiconductors theme","SOXX,SMH",5),
  ("Technology themes","Fintech/blockchain","ARKF,BLOK",6),
  ("Technology themes","Internet","FDN,PNQI",5),
  ("Disruptive","Innovation/disruption","ARKK,ARKW",6),
  ("Disruptive","Genomics","ARKG,IDNA",6),
  ("Disruptive","Space economy","UFO,ROKT,ARKX",6),
  ("Disruptive","Robotics & automation","BOTZ,ROBO",6),
  ("Demographics","Aging/longevity","-",4),
  ("Demographics","Millennials/Gen-Z","-",5),
  ("Energy transition","Clean energy","ICLN,TAN,FAN",6),
  ("Energy transition","Lithium/battery","LIT,BATT",6),
  ("Energy transition","Hydrogen","HYDR,HJEN",6),
  ("Energy transition","Uranium/nuclear","URA,URNM,NLR",6),
  ("Energy transition","EV/autonomous","DRIV,IDRV,KARS",6),
  ("Infrastructure theme","Reshoring/industrials","-",5),
  ("Infrastructure theme","Water","PHO,FIW,CGW",3),
  ("Infrastructure theme","Smart grid/5G","-",5),
  ("ESG/values","ESG-screened","ESGU,SUSA",4),
  ("ESG/values","Faith-based","-",4),
  ("ESG/values","Gender diversity","SHE",4),
  ("Commodity-linked equity","Gold miners","GDX,GDXJ",6),
  ("Commodity-linked equity","Agriculture/agribusiness","MOO,VEGI",5),
  ("Commodity-linked equity","Timber/forestry","WOOD,CUT",4),
  ("Defense/security","Defense","ITA,PPA,XAR",3),
  ("Cannabis","Cannabis","MSOS,MJ",6),
  ("Gaming/esports","Video games/esports","ESPO,HERO",6),
]:
    R(T,sub,sec,"Thematic ETF","-","Global",ex,"0-2%",risk,"Intraday","Theme-adoption bet; concentration risk")

# =====================================================================
# TIER 12 — REAL ESTATE (vehicles beyond GICS REIT sub-industries)
# =====================================================================
T = "Real Estate Vehicles"
R(T,"Listed REITs","Broad REIT","Diversified","US REIT index","Domestic","VNQ,SCHH,IYR","3-4%",4,"Intraday","Property cycle + rates")
R(T,"Listed REITs","Net lease","Triple-net","Net-lease REIT","Domestic","NETL","4-5%",3,"Intraday","Bond-like lease income")
R(T,"Mortgage REITs","Agency mREIT","Agency MBS levered","Agency mREIT","Domestic","REM,MORT","9-13%",6,"Intraday","Levered MBS carry; rate risk")
R(T,"Mortgage REITs","Commercial mREIT","CRE debt","Commercial mREIT","Domestic","-","8-11%",6,"Intraday","CRE credit + leverage")
R(T,"International RE","Global REIT","Ex-US property","Global REIT","Global","VNQI,RWX","3-4%",4,"Intraday","Foreign property + FX")
R(T,"Private RE","Non-traded REIT","Core/core-plus","Private real estate","Domestic","-","4-6%",5,"Periodic","Appraisal NAV; gates")
R(T,"Private RE","Value-add/opportunistic","LP funds","Private RE equity","Global","-","-",6,"Illiquid","Development/repositioning risk")
R(T,"Direct property","Residential","Rental/SFR/multifamily","Direct ownership","Domestic","-","4-8%",5,"Illiquid","Leverage + concentration")
R(T,"Direct property","Commercial","Office/retail/industrial","Direct ownership","Domestic","-","5-8%",5,"Illiquid","Tenant/vacancy risk")
R(T,"Direct property","Crowdfunding","Fractional RE","Platforms","Domestic","Fundrise*","4-8%",5,"Periodic","Platform/illiquidity risk")
R(T,"Land","Farmland","Agricultural land","Farmland","Domestic","FPI,LAND","2-4%",4,"Daily","Crop income + appreciation")
R(T,"Land","Timberland","Forest land","Timber","Domestic","WY,PCH","2-4%",4,"Daily","Biological growth + lumber price")
R(T,"Land","Raw/development land","Undeveloped","Land banking","Domestic","-","0%",5,"Illiquid","No income; entitlement risk")

# =====================================================================
# TIER 13 — INFRASTRUCTURE & MLPs
# =====================================================================
T = "Infrastructure"
R(T,"Listed infra","Core infrastructure","Global infra","Listed infra","Global","IGF,NFRA","2-4%",3,"Intraday","Regulated, inflation-linked")
R(T,"MLPs","Midstream MLP","Pipelines/storage","Energy MLP","Domestic","AMLP,MLPA,MLPX","6-8%",4,"Intraday","K-1; toll-road cash flow")
R(T,"Digital infra","Towers/data/fiber","Communications infra","Digital infra","Global","SRVR,DTCR","2-3%",4,"Intraday","Data growth")
R(T,"Renewable infra","Yieldcos","Wind/solar assets","Renewable infra","Global","-","4-6%",4,"Daily","Contracted power; rate sensitive")
R(T,"Transport infra","Toll roads/airports/ports","Concessions","Transport infra","Global","-","3-5%",3,"Daily","Traffic/usage volumes")
R(T,"Social infra","Schools/hospitals/PPP","Availability-based","Social infra","Global","-","4-6%",3,"Periodic","Government counterparty")
R(T,"Private infra","Unlisted infra","Brownfield/greenfield","Private infra","Global","-","5-8%",5,"Illiquid","Long-dated, illiquid")

# =====================================================================
# TIER 14 — COMMODITIES (full contract universe)
# =====================================================================
T = "Commodities"
R(T,"Broad","Diversified basket","Multi-commodity","Broad commodity","Global","DBC,PDBC,GSG,BCI","0%",4,"Intraday","Inflation/real-asset hedge")
# (SubClass, SubSector contract, Example, Risk)
COMMS = {
 "Precious metals": [("Gold","GLD,IAU,GLDM",3),("Silver","SLV,SIVR",4),("Platinum","PPLT",5),("Palladium","PALL",5),("Rhodium","-",6)],
 "Energy": [("WTI crude oil","USO,CL",5),("Brent crude oil","BNO",5),("Natural gas","UNG,BOIL",6),("Heating oil","UHN",5),("Gasoline (RBOB)","UGA",5),("Gasoil","-",5),("Ethanol","-",5),("Coal","-",5),("Uranium","URA,SRUUF",6),("Electricity","-",6),("Carbon allowances","KRBN,GRN",5)],
 "Base/industrial metals": [("Copper","CPER,COPX",5),("Aluminum","JJU",5),("Zinc","-",5),("Nickel","JJN",5),("Lead","-",5),("Tin","-",5),("Iron ore","-",5),("Cobalt","-",6),("Lithium","LIT",6),("Molybdenum","-",6),("Steel/rebar","SLX",5)],
 "Grains & oilseeds": [("Corn","CORN",5),("Wheat (SRW)","WEAT",5),("Wheat (HRW)","-",5),("Soybeans","SOYB",5),("Soybean meal","-",5),("Soybean oil","-",5),("Rough rice","-",5),("Oats","-",5),("Canola/rapeseed","-",5),("Palm oil","-",5)],
 "Softs": [("Coffee (Arabica)","JO",6),("Coffee (Robusta)","-",6),("Sugar #11","CANE",6),("Cocoa","NIB",6),("Cotton","BAL",6),("Orange juice","-",6),("Lumber","WOOD",5),("Natural rubber","-",6)],
 "Livestock": [("Live cattle","COW*",5),("Feeder cattle","-",5),("Lean hogs","-",5),("Class III milk","-",5)],
}
for sub, items in COMMS.items():
    for contract, ex, risk in items:
        R(T,sub,contract,"Futures/physical","-","Global",ex,"0%",risk,"Intraday","Spot price + roll yield; storage")
R(T,"Physical","Allocated metals","Vaulted bullion","Physical holdings","Global","-","0%",3,"Periodic","Storage/insurance cost")
R(T,"Physical","Numismatic/bullion coins","Coins/bars","Physical coins","Global","-","0%",4,"Periodic","Premium over spot")

# =====================================================================
# TIER 15 — ALTERNATIVES / HEDGE STRATEGIES
# =====================================================================
T = "Alternatives"
R(T,"Liquid alts","Managed futures","Trend following","CTA","Global","DBMF,KMLM,CTA","0%",4,"Daily","Crisis-alpha; trend diversifier")
R(T,"Liquid alts","Global macro","Discretionary/systematic","Macro","Global","-","0%",4,"Daily","Top-down cross-asset bets")
R(T,"Liquid alts","Market neutral","Equity long/short","Market neutral","Global","BTAL","2-4%",3,"Daily","Low beta; spread capture")
R(T,"Liquid alts","Merger arbitrage","Event-driven","Merger arb","Global","MNA,MERFX","3-5%",3,"Daily","Deal-spread capture")
R(T,"Liquid alts","Convertible arbitrage","Vol arb","Convert arb","Global","-","3-5%",4,"Daily","Long convert/short equity")
R(T,"Liquid alts","Long/short equity","Hedged equity","L/S equity","Global","-","0-2%",4,"Daily","Net long with hedges")
R(T,"Liquid alts","Event-driven","Special situations","Event-driven","Global","-","varies",4,"Daily","Catalyst-driven")
R(T,"Liquid alts","Fixed income arb","Relative value","FI arb","Global","-","3-6%",4,"Daily","Curve/basis trades")
R(T,"Liquid alts","Multi-strategy","Diversified alts","Multi-strat","Global","QAI","2-4%",3,"Daily","Blended alt sleeves")
R(T,"Liquid alts","Risk parity","Levered diversified","Risk parity","Global","RPAR,UPAR","2-3%",4,"Daily","Vol-balanced multi-asset")
R(T,"Liquid alts","Return stacking","Capital efficient","Stacked beta+alt","Global","RSST,RSBT","1-3%",4,"Daily","Overlay alts on core beta")
R(T,"Hedge funds","Private hedge funds","LP structures","HF strategies","Global","-","varies",5,"Periodic","Lockups; 2/20; manager risk")
R(T,"Hedge funds","Fund of funds","Diversified HF","FoF","Global","-","varies",5,"Periodic","Layered fees; diversification")
R(T,"Insurance-linked","Catastrophe bonds","Reinsurance risk","Cat bonds","Global","-","8-12%",5,"Periodic","Uncorrelated; tail event loss")
R(T,"Insurance-linked","Life settlements","Longevity","Life settlements","Domestic","-","8-12%",5,"Illiquid","Longevity/mortality risk")
R(T,"Royalties","Music royalties","Catalog income","Music IP","Global","SONG*","6-9%",4,"Periodic","Streaming income streams")
R(T,"Royalties","Pharma royalties","Drug royalties","Pharma IP","Global","RPRX","6-9%",4,"Daily","Patent-life cash flows")
R(T,"Royalties","Mineral/oil royalties","Resource royalties","Royalty trusts","Domestic","-","8-12%",5,"Daily","Commodity + depletion")
R(T,"Litigation finance","Legal funding","Case portfolios","Litigation finance","Global","-","high",6,"Illiquid","Binary case outcomes")
R(T,"Equipment leasing","Aircraft/rail/marine","Leasing","Equipment finance","Global","-","6-10%",5,"Illiquid","Residual value + lessee credit")

# =====================================================================
# TIER 16 — PRIVATE EQUITY / VENTURE
# =====================================================================
T = "Private Equity"
R(T,"Listed PE","PE proxies","Listed sponsors","PE-linked equity","Domestic","PSP,KKR,BX,APO","2-4%",5,"Intraday","Public proxy for PE")
R(T,"Buyout","LBO funds","Large/mid buyout","Buyout","Global","-","-",6,"Illiquid","Leverage + operational alpha")
R(T,"Growth equity","Growth funds","Pre-IPO growth","Growth equity","Global","-","-",6,"Illiquid","Late-stage scaling")
R(T,"Venture capital","VC funds","Seed/early/late","Venture","Global","-","-",6,"Illiquid","Power-law return; J-curve")
R(T,"Venture capital","Angel/direct","Single startups","Angel","Global","-","-",6,"Illiquid","Total-loss risk")
R(T,"Secondaries","PE secondaries","LP stakes","Secondaries","Global","-","-",5,"Illiquid","Discounted NAV entry")
R(T,"Co-investment","Direct co-invest","Deal-by-deal","Co-investment","Global","-","-",6,"Illiquid","Concentrated single-deal")
R(T,"Interval/evergreen","Evergreen PE","Retail-access PE","Evergreen","Global","-","-",5,"Periodic","Quarterly gates; NAV marks")
R(T,"Pre-IPO","Late-stage private","Secondary shares","Pre-IPO","Global","-","-",6,"Illiquid","Liquidity timing risk")

# =====================================================================
# TIER 17 — VOLATILITY & OPTIONS / DERIVATIVES / STRUCTURED PRODUCTS
# =====================================================================
T = "Derivatives & Structured"
R(T,"Long vol","VIX futures","Short-term VIX","Long volatility","Domestic","VIXY,VXX,UVXY","0%",6,"Intraday","Hedge; severe roll decay")
R(T,"Tail risk","Tail hedge","Convexity","Tail risk","Domestic","TAIL","0%",5,"Intraday","Crash insurance; bleeds")
R(T,"Option income","Covered call","Buy-write","Option income","Domestic","JEPI,JEPQ,QYLD,XYLD","7-12%",4,"Intraday","Caps upside for income")
R(T,"Option income","Put-write","Cash-secured puts","Put income","Domestic","PUTW","5-8%",4,"Intraday","Sells downside insurance")
R(T,"Defined outcome","Buffered/defined","Structured ETF","Buffer/floor","Domestic","BUFR,PBP","varies",3,"Intraday","Caps gains for downside buffer")
R(T,"Short vol","Inverse VIX","Short volatility","Short vol","Domestic","SVIX,SVXY","0%",6,"Intraday","Sells vol; blow-up risk")
R(T,"Listed options","Equity/index options","Calls/puts/spreads","Options","Domestic","SPX,SPY opts","-",6,"Intraday","Leverage; time decay")
R(T,"Futures","Index/rate/commodity futures","Exchange futures","Futures","Global","ES,ZN,GC","-",5,"Intraday","Leverage; margin")
R(T,"Swaps","Interest rate swaps","IRS","OTC swaps","Global","-","-",4,"Periodic","Counterparty/rate risk")
R(T,"Swaps","Credit default swaps","CDS/CDX","Credit swaps","Global","-","-",5,"Periodic","Credit-event protection")
R(T,"Swaps","Total return swaps","TRS","Synthetic exposure","Global","-","-",5,"Periodic","Financing + counterparty")
R(T,"Swaps","Variance/vol swaps","Vol swaps","OTC vol","Global","-","-",6,"Periodic","Pure volatility exposure")
R(T,"Structured notes","Autocallables","Income notes","Autocall","Global","-","6-12%",4,"Periodic","Issuer credit; capped/contingent")
R(T,"Structured notes","Principal-protected","PPN/market-linked CD","Protected notes","Global","-","varies",3,"Periodic","Issuer credit; opportunity cost")
R(T,"Structured notes","Reverse convertibles","Yield notes","Rev converts","Global","-","8-15%",5,"Periodic","Downside equity risk")
R(T,"Warrants/rights","Warrants","Long-dated calls","Warrants","Global","-","0%",6,"Intraday","Dilution; expiry")
R(T,"Leveraged/inverse","Geared ETFs","2x/3x/-1x","L&I ETPs","Domestic","TQQQ,SQQQ,UPRO","0%",6,"Intraday","Daily reset decay")

# =====================================================================
# TIER 18 — DIGITAL ASSETS
# =====================================================================
T = "Digital Assets"
R(T,"Crypto","Bitcoin","Spot BTC","Store of value","Global","IBIT,FBTC,GBTC","0%",6,"Intraday","Digital gold thesis; high vol")
R(T,"Crypto","Ethereum","Spot ETH","Smart-contract L1","Global","ETHA,ETHE","0-4%*",6,"Intraday","Platform bet; *staking yield")
R(T,"Crypto","Large-cap L1","SOL/ADA/AVAX/DOT","Layer-1 alts","Global","-","0%",6,"Intraday","High beta to BTC")
R(T,"Crypto","Layer-2/scaling","Rollups/sidechains","Layer-2","Global","-","0%",6,"Intraday","Scaling adoption bet")
R(T,"Crypto","DeFi tokens","DEX/lending tokens","DeFi","Global","-","0%",6,"Intraday","Protocol risk")
R(T,"Crypto","Stablecoin yield","Lending/staking","DeFi yield","Global","-","4-10%",5,"Intraday","Smart-contract/counterparty risk")
R(T,"Crypto","Staking/PoS","Validator yield","Staking","Global","-","3-8%",5,"Periodic","Lockup/slashing risk")
R(T,"Crypto","Mining","Hashpower","Mining","Global","WGMI","0%",6,"Daily","Hardware/energy economics")
R(T,"Crypto","Crypto equity","Miners/exchanges/treasuries","Crypto-linked stocks","Global","COIN,MSTR,WGMI","0%",6,"Intraday","Levered crypto proxy")
R(T,"Crypto","Futures/derivatives","Perps/futures","Crypto derivs","Global","BITO","0%",6,"Intraday","Leverage; funding")
R(T,"NFTs","Digital collectibles","Art/PFP/utility","NFTs","Global","-","0%",6,"Illiquid","Highly illiquid; fad risk")
R(T,"Tokenized","RWA tokens","Tokenized T-bills/credit","Tokenized RWA","Global","-","4-8%",4,"Intraday","On-chain real assets")
R(T,"Tokenized","Tokenized equities/funds","On-chain securities","Tokenized funds","Global","-","varies",4,"Intraday","Regulatory/custody risk")

# =====================================================================
# TIER 19 — CURRENCIES / FX
# =====================================================================
T = "Currencies / FX"
R(T,"USD","Dollar index","DXY","Long/short USD","Global","UUP,UDN","0-4%",3,"Intraday","Dollar direction")
FX_PAIRS = {
 "Major pairs": [("EUR/USD","FXE"),("USD/JPY","FXY"),("GBP/USD","FXB"),("USD/CHF","FXF"),("USD/CAD","FXC"),("AUD/USD","FXA"),("NZD/USD","BNZ*")],
 "Cross pairs": [("EUR/GBP","-"),("EUR/JPY","-"),("GBP/JPY","-"),("EUR/CHF","-"),("AUD/JPY","-"),("CHF/JPY","-")],
 "EM pairs": [("USD/CNH","CYB"),("USD/INR","-"),("USD/BRL","BZF"),("USD/MXN","-"),("USD/ZAR","-"),("USD/TRY","-"),("USD/KRW","-"),("USD/SGD","-"),("USD/HKD","-"),("USD/IDR","-"),("USD/THB","-"),("USD/PLN","-")],
 "Scandies/commodity FX": [("USD/NOK","-"),("USD/SEK","-"),("USD/CAD (oil)","FXC"),("AUD (metals)","FXA")],
}
for sub, pairs in FX_PAIRS.items():
    for pair, ex in pairs:
        R(T,"FX pairs",sub,pair,"-","Global",ex,"0-5%",4,"Intraday","Rate differential + spot")
R(T,"Carry","High-yield carry","Carry basket","Carry","Global","-","3-6%",4,"Daily","Carry; crash risk")
R(T,"Gold as currency","Monetary metal","Anti-fiat","Bullion","Global","GLD,IAU","0%",3,"Intraday","Currency-debasement hedge")

# =====================================================================
# TIER 20 — INSURANCE / ANNUITIES / RETIREMENT VEHICLES
# =====================================================================
T = "Insurance & Annuities"
R(T,"Annuities","Fixed annuity","MYGA","Guaranteed rate","Domestic","-","4-6%",2,"Illiquid","Insurer credit; surrender charges")
R(T,"Annuities","Fixed-indexed annuity","FIA","Index-linked w/floor","Domestic","-","varies",3,"Illiquid","Capped upside; principal floor")
R(T,"Annuities","Variable annuity","VA","Subaccounts","Domestic","-","varies",4,"Illiquid","Market risk + fees + riders")
R(T,"Annuities","Immediate/deferred income","SPIA/DIA/QLAC","Lifetime income","Domestic","-","varies",2,"Illiquid","Longevity pooling; illiquid")
R(T,"Annuities","Buffered/RILA","Registered index-linked","Buffer annuity","Domestic","-","varies",3,"Illiquid","Partial downside buffer")
R(T,"Life insurance","Whole life","Cash value","Permanent life","Domestic","-","2-4%",2,"Illiquid","Guaranteed cash value + dividends")
R(T,"Life insurance","Universal/IUL","Flexible permanent","UL/IUL","Domestic","-","varies",3,"Illiquid","Crediting + cost of insurance")
R(T,"Structured settlements","Settlement annuities","Secondary market","Settlements","Domestic","-","4-6%",3,"Illiquid","Discounted future payments")
R(T,"Pension","Pension buy-in/out","Risk transfer","Pension","Domestic","-","-",2,"Illiquid","Institutional longevity transfer")

# =====================================================================
# TIER 21 — COLLECTIBLES / ESOTERIC REAL ASSETS / NATURAL RESOURCE RIGHTS
# =====================================================================
T = "Collectibles & Esoteric"
ESOTERIC = [
 ("Collectibles","Fine art","Blue-chip art","Masterworks*",6,"Illiquid","Taste-driven; illiquid"),
 ("Collectibles","Fine wine","Vintage wine","Vinovest*",5,"Illiquid","Storage; provenance"),
 ("Collectibles","Whiskey/spirits casks","Maturing casks","-",5,"Illiquid","Maturation/storage"),
 ("Collectibles","Watches","Luxury timepieces","-",6,"Illiquid","Brand/condition dependent"),
 ("Collectibles","Jewelry/gemstones","Diamonds/gems","-",6,"Illiquid","Opaque pricing"),
 ("Collectibles","Classic/collector cars","Vintage autos","-",6,"Illiquid","Maintenance; taste"),
 ("Collectibles","Trading cards","Sports/TCG","-",6,"Illiquid","Fad risk"),
 ("Collectibles","Sports memorabilia","Memorabilia","-",6,"Illiquid","Authentication risk"),
 ("Collectibles","Stamps (philately)","Rare stamps","-",6,"Illiquid","Niche market"),
 ("Collectibles","Coins (numismatics)","Rare coins","-",5,"Illiquid","Grading/premium"),
 ("Collectibles","Comic books","Vintage comics","-",6,"Illiquid","Condition/grading"),
 ("Collectibles","Sneakers/streetwear","Resale market","-",6,"Illiquid","Trend-driven"),
 ("Collectibles","Handbags","Luxury bags","-",6,"Illiquid","Brand/condition"),
 ("Collectibles","Musical instruments","Rare instruments","-",6,"Illiquid","Specialist market"),
 ("Collectibles","Books/manuscripts","Rare books","-",6,"Illiquid","Provenance"),
 ("Digital esoteric","Domain names","Premium domains","-",6,"Illiquid","Speculative; carrying cost"),
 ("Natural resource rights","Mineral rights","Subsurface","-",5,"Illiquid","Commodity + depletion"),
 ("Natural resource rights","Oil & gas royalties","Royalty interests","-",5,"Daily","Production decline risk"),
 ("Natural resource rights","Water rights","Allocations","-",5,"Illiquid","Regulatory/scarcity"),
 ("Natural resource rights","Carbon credits","Voluntary/compliance","KRBN",5,"Daily","Policy-dependent"),
 ("Natural resource rights","Timber rights","Cutting rights","-",4,"Illiquid","Biological growth"),
 ("Operating real assets","Vessels/shipping","Tankers/bulkers","-",6,"Illiquid","Freight-rate cycle"),
 ("Operating real assets","Aircraft leasing","Aviation assets","-",5,"Illiquid","Residual value/lessee"),
 ("Operating real assets","Rail cars/containers","Transport leasing","-",5,"Illiquid","Utilization rates"),
 ("Operating real assets","Solar/wind farms","Power assets","-",4,"Illiquid","Contracted offtake"),
 ("Operating real assets","Data center capacity","Compute assets","-",5,"Illiquid","Tenant demand"),
 ("Operating real assets","Cell sites/billboards","Site leases","-",4,"Illiquid","Lease cash flows"),
 ("IP","Patents","Patent portfolios","-",6,"Illiquid","Enforcement/expiry"),
 ("IP","Film/TV royalties","Content libraries","-",5,"Periodic","Distribution risk"),
 ("IP","Trademarks/brands","Brand licensing","-",5,"Illiquid","Brand value"),
]
for sub, sec, subsec, ex, risk, liq, note in ESOTERIC:
    R(T,sub,sec,subsec,"-","Global",ex,"varies",risk,liq,note)

# =====================================================================
# ENRICH: primary ticker + real 5Y metrics (yield / Sharpe / return)
# Metrics are loaded from metrics.json (populated from public aggregators
# via web search; figures are dated and provider-sourced, NOT computed here
# and NOT advice). Sleeves with no liquid proxy show N/A.
# =====================================================================
import json
metrics_path = os.path.join(OUT_DIR, "metrics.json")
try:
    with open(metrics_path) as f:
        METRICS = {k.upper(): v for k, v in json.load(f).items()}
except FileNotFoundError:
    METRICS = {}

def get_primary(ex):
    if not ex or ex == "-":
        return ""
    first = ex.split(",")[0].strip().lstrip("~").rstrip("*").strip()
    first = first.split()[0] if first else ""
    return first

ENRICHED = []
for row in ROWS:
    pt = get_primary(row[6])
    m = METRICS.get(pt.upper()) if pt else None
    if m:
        y = m.get("yield", "N/A"); s = m.get("sharpe", "N/A")
        c = m.get("cagr", "N/A"); src = m.get("src", "")
    else:
        y = s = c = "N/A"; src = ("no liquid proxy" if not pt else "not retrieved")
    ENRICHED.append(tuple(list(row) + [pt or "—", y, s, c, src]))

COLUMNS = COLUMNS + ["Primary Ticker", "5Y Yield", "5Y Sharpe", "5Y Return (CAGR)", "Metric Source"]
ROWS = ENRICHED
_metrics_filled = sum(1 for r in ROWS if r[12] != "N/A")
print(f"Metrics filled (real): {_metrics_filled} of {len(ROWS)} sleeves; "
      f"unique tickers in metrics.json: {len(METRICS)}")

# =====================================================================
# WRITE CSV
# =====================================================================
csv_path = os.path.join(OUT_DIR, "asset_universe.csv")
with open(csv_path, "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(COLUMNS)
    for row in ROWS:
        w.writerow(row)

# =====================================================================
# WRITE EXCEL
# =====================================================================
wb = Workbook()
header_fill = PatternFill("solid", fgColor="1F3864")
header_font = Font(color="FFFFFF", bold=True, size=11)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
wrap = Alignment(vertical="top", wrap_text=True)
risk_colors = {1:"C6EFCE",2:"D9EAD3",3:"FFF2CC",4:"FCE5CD",5:"F4CCCC",6:"EA9999"}
WIDTHS = [26, 24, 26, 30, 32, 16, 26, 12, 9, 12, 44, 13, 11, 10, 14, 26]
CENTER_COLS = {8, 9, 10, 12, 13, 14, 15}

def style_sheet(ws, rows, with_tracking=False):
    headers = COLUMNS + (["Hold?","Position $","Target %","Notes/Action"] if with_tracking else [])
    ws.append(headers)
    for c, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = center; cell.border = border
    for row in rows:
        ws.append(list(row) + (["","","",""] if with_tracking else []))
    widths = WIDTHS + ([8,12,9,28] if with_tracking else [])
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = center if c in CENTER_COLS else wrap
        rt = ws.cell(row=r, column=9).value
        try:
            rt = int(rt)
            ws.cell(row=r, column=9).fill = PatternFill("solid", fgColor=risk_colors.get(rt,"FFFFFF"))
        except (TypeError, ValueError):
            pass
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

ws_master = wb.active
ws_master.title = "MASTER (all assets)"
style_sheet(ws_master, ROWS, with_tracking=True)

tiers = []
for row in ROWS:
    if row[0] not in tiers:
        tiers.append(row[0])
for tier in tiers:
    safe = tier.replace("/", "-")[:31]
    ws = wb.create_sheet(title=safe)
    style_sheet(ws, [r for r in ROWS if r[0] == tier], with_tracking=False)

ws_leg = wb.create_sheet(title="LEGEND", index=1)
legend = [
 ["Asset Universe — Exhaustive Tracking Taxonomy",""],
 ["",""],
 ["Coverage", f"{len(ROWS)} sleeves across {len(tiers)} tiers"],
 ["Hierarchy","Tier > Asset Class > Sub-Class > Sector/Strategy > Sub-Sector"],
 ["Equity sectors","Full GICS 2023: 11 sectors > 25 groups > 74 industries > 163 sub-industries (global)"],
 ["Geography","Domestic (US) / Developed ex-US / EM / Frontier / Global"],
 ["",""],
 ["Risk Tier scale",""],
 ["1","Capital preservation (cash, T-bills, AAA CLO, agency MBS)"],
 ["2","Low risk (short IG, TIPS, mezz CLO debt, MYGA)"],
 ["3","Income + some protection (preferreds, utilities, gold)"],
 ["4","Moderate (broad equity, HY, REITs, commodities, CLO BB)"],
 ["5","Elevated (small cap, EM, BDCs, sector bets, loans)"],
 ["6","Max beta / first loss (CLO equity, crypto, VIX, biotech, PE/VC)"],
 ["",""],
 ["Liquidity","Intraday / Daily / Periodic (gated) / Illiquid (lockup)"],
 ["Typ. Yield","Rough illustrative income range; TE = tax-exempt; *=conditional"],
 ["",""],
 ["DISCLAIMER","Illustrative taxonomy for tracking/organization only."],
 ["","Tickers are examples to anchor categories, NOT recommendations."],
 ["","Yields/risk tiers are approximate as of generation, not advice."],
 ["",""],
 ["The original CLO ladder maps to:",""],
 ["1 Safest — AAA CLO debt","Securitized Credit > CLO > CLO debt > AAA (JAAA)"],
 ["2 Sweet spot — Mezz CLO debt","Securitized Credit > CLO > CLO debt > BBB (JBBB,CLOZ)"],
 ["3 Income+protection — CLO-CEF pfd","Preferred & Hybrid > CLO-CEF preferreds (OXLCN/I/Z,ECC-D)"],
 ["4 Raw asset — Senior loans","Loans > Senior loans (BKLN,SRLN)"],
 ["5 Direct private credit — BDC","Private Credit > BDC (ARCC,BIZD)"],
 ["6 Max beta — CLO equity","Securitized Credit > CLO > CLO equity (ECC,OXLC)"],
]
for r in legend:
    ws_leg.append(r)
ws_leg.column_dimensions["A"].width = 40
ws_leg.column_dimensions["B"].width = 78
ws_leg["A1"].font = Font(bold=True, size=14, color="1F3864")
for r in range(1, ws_leg.max_row + 1):
    a = ws_leg.cell(row=r, column=1)
    if a.value in ("Risk Tier scale","DISCLAIMER","The original CLO ladder maps to:","Hierarchy","Coverage","Equity sectors"):
        a.font = Font(bold=True)
    ws_leg.cell(row=r, column=2).alignment = wrap

xlsx_path = os.path.join(OUT_DIR, "asset_universe.xlsx")
wb.save(xlsx_path)

print(f"Rows: {len(ROWS)}")
print(f"Tiers: {len(tiers)}")
for t in tiers:
    print(f"  {t}: {sum(1 for r in ROWS if r[0]==t)}")
print(f"CSV : {csv_path}")
print(f"XLSX: {xlsx_path}")
